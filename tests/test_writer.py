"""测试 Excel 写入工具"""

import os
import pytest
import openpyxl

from tools.writer import create_workbook, write_cells, apply_number_format

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
OUTPUT = os.path.join(FIXTURES, "test_output.xlsx")


@pytest.fixture(autouse=True)
def cleanup():
    os.makedirs(FIXTURES, exist_ok=True)
    yield
    if os.path.exists(OUTPUT):
        os.remove(OUTPUT)


def test_create_workbook():
    create_workbook(OUTPUT, {
        "Sheet1": [
            ["Name", "Age"],
            ["Alice", 30],
            ["Bob", 25],
        ]
    })
    assert os.path.exists(OUTPUT)

    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == 30
    wb.close()


def test_write_cells_with_formula():
    create_workbook(OUTPUT, {
        "Data": [
            ["A", "B", "Sum"],
            [10, 20, None],
        ]
    })

    write_cells(OUTPUT, "Data", [
        {"cell": "C2", "value": "=A2+B2", "style": "formula"},
    ])

    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb["Data"]
    assert ws["C2"].value == "=A2+B2"
    wb.close()


def test_write_cells_with_styles():
    create_workbook(OUTPUT, {"Sheet1": [["Value"]]})

    write_cells(OUTPUT, "Sheet1", [
        {"cell": "A2", "value": 1000, "style": "input"},
        {"cell": "A3", "value": 0.05, "style": "assumption"},
    ])

    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    # 蓝色输入 (openpyxl 存储时带 alpha 前缀 "00")
    assert ws["A2"].font.color.rgb.endswith("0000FF")
    # 黄色背景假设
    assert ws["A3"].fill.start_color.rgb.endswith("FFFF00")
    wb.close()


def test_apply_number_format():
    create_workbook(OUTPUT, {
        "Sheet1": [
            ["Amount"],
            [1234.56],
            [-789.01],
        ]
    })

    apply_number_format(OUTPUT, "Sheet1", [
        {"range": "A2:A3", "format": '#,##0.00;(#,##0.00);"-"'},
    ])

    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    assert ws["A2"].number_format == '#,##0.00;(#,##0.00);"-"'
    wb.close()
