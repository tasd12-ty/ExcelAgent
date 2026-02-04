"""Excel 公式重算与验证脚本。

用法: python scripts/recalc.py <excel_file>

功能:
- 打开 Excel 文件
- 检测所有公式
- 报告潜在的公式错误 (#REF!, #DIV/0!, #VALUE! 等)
"""

import sys

import openpyxl


ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#N/A", "#NUM!"}


def recalc(file_path: str) -> dict:
    """检查 Excel 文件中的公式和错误。"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)

    results = {"formulas": 0, "errors": []}

    for sheet_name in wb.sheetnames:
        ws_val = wb[sheet_name]
        ws_formula = wb_formulas[sheet_name]

        for row_val, row_formula in zip(ws_val.iter_rows(), ws_formula.iter_rows()):
            for cell_val, cell_formula in zip(row_val, row_formula):
                if isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                    results["formulas"] += 1
                    val_str = str(cell_val.value) if cell_val.value else ""
                    if val_str in ERROR_VALUES:
                        results["errors"].append({
                            "sheet": sheet_name,
                            "cell": cell_formula.coordinate,
                            "formula": cell_formula.value,
                            "error": val_str,
                        })

    wb.close()
    wb_formulas.close()
    return results


def main():
    if len(sys.argv) < 2:
        print("用法: python recalc.py <excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    results = recalc(file_path)

    print(f"公式总数: {results['formulas']}")
    if results["errors"]:
        print(f"发现 {len(results['errors'])} 个错误:")
        for err in results["errors"]:
            print(f"  [{err['sheet']}] {err['cell']}: {err['formula']} -> {err['error']}")
    else:
        print("未发现公式错误。")


if __name__ == "__main__":
    main()
