"""Excel 格式化工具"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def auto_fit_columns(file_path: str, sheet_name: str) -> str:
    """自动调整列宽以适应内容。"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(file_path)
    wb.close()
    return file_path


def apply_header_style(
    file_path: str,
    sheet_name: str,
    header_row: int = 1,
    bg_color: str = "4472C4",
    font_color: str = "FFFFFF",
) -> str:
    """给表头行应用样式。"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    font = Font(color=font_color, bold=True)
    alignment = Alignment(horizontal="center")

    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    wb.save(file_path)
    wb.close()
    return file_path


def add_table_borders(file_path: str, sheet_name: str, cell_range: str) -> str:
    """给指定区域添加边框。

    Args:
        cell_range: e.g. "A1:D10"
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

    wb.save(file_path)
    wb.close()
    return file_path
