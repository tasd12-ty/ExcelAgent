"""Excel 文件读取工具"""

import os
from typing import Any

import openpyxl
import pandas as pd


def read_excel(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """读取 Excel 文件，返回结构化信息。

    Returns:
        {
            "file": 文件名,
            "sheets": [sheet名称列表],
            "active_sheet": 当前活动sheet,
            "data": {sheet_name: DataFrame},
            "shape": {sheet_name: (rows, cols)},
        }
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    active = wb.active.title if wb.active else sheets[0]

    target_sheets = [sheet_name] if sheet_name and sheet_name in sheets else sheets

    data = {}
    shape = {}
    for name in target_sheets:
        df = pd.read_excel(file_path, sheet_name=name, engine="openpyxl")
        data[name] = df
        shape[name] = df.shape

    wb.close()

    return {
        "file": os.path.basename(file_path),
        "sheets": sheets,
        "active_sheet": active,
        "data": data,
        "shape": shape,
    }


def read_excel_formulas(file_path: str, sheet_name: str | None = None) -> dict[str, list]:
    """读取 Excel 中的公式（非计算值）。

    Returns:
        {sheet_name: [{"cell": "A1", "formula": "=SUM(B1:B10)"}, ...]}
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames

    formulas = {}
    for name in sheets:
        ws = wb[name]
        sheet_formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value,
                    })
        formulas[name] = sheet_formulas

    wb.close()
    return formulas


def get_summary(file_path: str) -> str:
    """生成 Excel 文件的文本摘要，适合发送给 LLM。"""
    info = read_excel(file_path)
    lines = [f"文件: {info['file']}", f"工作表: {', '.join(info['sheets'])}"]

    for sheet_name, df in info["data"].items():
        rows, cols = info["shape"][sheet_name]
        lines.append(f"\n--- {sheet_name} ({rows}行 x {cols}列) ---")
        lines.append(f"列名: {', '.join(str(c) for c in df.columns)}")
        if rows > 0:
            lines.append("前5行数据:")
            lines.append(df.head().to_string(index=False))

        # 数值列基本统计
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if numeric_cols:
            lines.append(f"\n数值列统计 ({', '.join(numeric_cols)}):")
            lines.append(df[numeric_cols].describe().to_string())

    return "\n".join(lines)
