"""Excel 文件写入工具"""

import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers


# 金融模型色彩规范
STYLE_INPUT = Font(color="0000FF")       # 蓝色：硬编码输入
STYLE_FORMULA = Font(color="000000")     # 黑色：公式
STYLE_LINK_INT = Font(color="008000")    # 绿色：内部链接
STYLE_LINK_EXT = Font(color="FF0000")    # 红色：外部链接
FILL_ASSUMPTION = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色背景：关键假设


def create_workbook(file_path: str, sheets: dict[str, list[list]]) -> str:
    """创建新 Excel 文件。

    Args:
        file_path: 输出路径
        sheets: {sheet_name: [[row1_data], [row2_data], ...]}

    Returns:
        创建的文件路径
    """
    wb = openpyxl.Workbook()
    first = True

    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row_data in rows:
            ws.append(row_data)

    wb.save(file_path)
    wb.close()
    return file_path


def write_cells(file_path: str, sheet_name: str, cells: list[dict]) -> str:
    """写入单元格数据。

    Args:
        cells: [{"cell": "A1", "value": xxx, "style": "input|formula|assumption"}, ...]
    """
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    for item in cells:
        cell = ws[item["cell"]]
        cell.value = item["value"]

        style = item.get("style", "")
        if style == "input":
            cell.font = STYLE_INPUT
        elif style == "formula":
            cell.font = STYLE_FORMULA
        elif style == "assumption":
            cell.font = STYLE_INPUT
            cell.fill = FILL_ASSUMPTION

    wb.save(file_path)
    wb.close()
    return file_path


def set_column_widths(file_path: str, sheet_name: str, widths: dict[str, int]) -> str:
    """设置列宽。

    Args:
        widths: {"A": 15, "B": 20, ...}
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(file_path)
    wb.close()
    return file_path


def apply_number_format(file_path: str, sheet_name: str, formats: list[dict]) -> str:
    """应用数字格式。

    Args:
        formats: [{"range": "B2:B10", "format": "#,##0.00"}, ...]

    常用格式:
        - 货币: '#,##0.00'
        - 百分比: '0.0%'
        - 负数括号: '#,##0.00;(#,##0.00)'
        - 零显示为横线: '#,##0.00;(#,##0.00);"-"'
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    for fmt in formats:
        for row in ws[fmt["range"]]:
            for cell in row:
                cell.number_format = fmt["format"]

    wb.save(file_path)
    wb.close()
    return file_path
